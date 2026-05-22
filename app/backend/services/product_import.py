import csv
import io
import logging
from typing import Dict, List, Optional, Any, Tuple

from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger(__name__)

# Dynamic field registry: maps model field name to metadata
PRODUCT_FIELD_REGISTRY: Dict[str, Dict[str, Any]] = {
    "name": {"required": True, "label": "name", "example": "Ocean Breeze"},
    "brand": {"required": True, "label": "brand", "example": "PlaceholderBrand1"},
    "category": {"required": False, "label": "category", "example": "Hanging Air Freshener"},
    "price": {"required": True, "label": "price", "example": "12.99"},
    "volumes": {"required": False, "label": "volumes", "example": "0, 10, 20, 30"},
    "image": {"required": False, "label": "image", "example": "https://example.com/image.jpg"},
    "description": {"required": False, "label": "description", "example": "Fresh ocean scent"},
    "instagram_url": {"required": False, "label": "instagram_url", "example": "https://instagram.com/brand"},
    "refillable": {"required": False, "label": "refillable", "example": "TRUE/FALSE"},
    "is_new": {"required": False, "label": "is_new", "example": "TRUE/FALSE"},
    "is_featured": {"required": False, "label": "is_featured", "example": "TRUE/FALSE"},
}

# Alias mapping: common variations → canonical field name
FIELD_ALIASES = {
    "product name": "name",
    "brand name": "brand",
    "type": "category",
    "product type": "category",
    "cost": "price",
    "price_usd": "price",
    "volume": "volumes",
    "sizes": "volumes",
    "image_url": "image",
    "image url": "image",
    "photo": "image",
    "desc": "description",
    "description": "description",
    "instagram": "instagram_url",
    "instagram url": "instagram_url",
    "instagramurl": "instagram_url",
    "can refill": "refillable",
    "new": "is_new",
    "new product": "is_new",
    "is new": "is_new",
    "featured": "is_featured",
    "is featured": "is_featured",
    "bestseller": "is_featured",
}

# Boolean true/false string representations
BOOL_TRUE = {"true", "yes", "1", "да", "y"}
BOOL_FALSE = {"false", "no", "0", "нет", "n"}


def _resolve_field_name(header: str) -> Optional[str]:
    """Map a raw file header to a canonical field name."""
    normalized = header.strip().lower().replace(" ", "_").replace("-", "_")
    if normalized in PRODUCT_FIELD_REGISTRY:
        return normalized
    if normalized in FIELD_ALIASES:
        return FIELD_ALIASES[normalized]
    return None


def _parse_bool(val: str) -> Optional[bool]:
    """Parse a boolean from various string representations."""
    if val.strip().lower() in BOOL_TRUE:
        return True
    if val.strip().lower() in BOOL_FALSE:
        return False
    return None


def _parse_price(val: str) -> Optional[float]:
    """Parse price, handling comma as decimal separator."""
    try:
        cleaned = val.strip().replace(" ", "").replace(",", ".")
        return float(cleaned)
    except (ValueError, AttributeError):
        return None


def _parse_volumes(val: str) -> Optional[str]:
    """Volumes stored as comma-separated string."""
    return val.strip() or None


def _validate_row(row_data: Dict[str, Any]) -> List[Dict[str, str]]:
    """Validate a single row. Returns list of errors."""
    errors = []
    name = row_data.get("name")
    brand = row_data.get("brand")
    price = row_data.get("price")
    
    if not name or str(name).strip() == "":
        errors.append({"field": "name", "message": "Обязательное поле"})
    if not brand or str(brand).strip() == "":
        errors.append({"field": "brand", "message": "Обязательное поле"})
    if price is not None:
        try:
            p = float(price)
            if p < 0:
                errors.append({"field": "price", "message": "Цена не может быть отрицательной"})
        except (ValueError, TypeError):
            errors.append({"field": "price", "message": "Должно быть число"})
    return errors


def _parse_row_to_product(
    row: Dict[str, str], headers: List[str]
) -> Dict[str, Any]:
    """Convert a raw CSV/Excel row dict to Product creation data."""
    result = {}
    for header in headers:
        field = _resolve_field_name(header)
        if field is None:
            continue
        raw_val = row.get(header, "").strip()
        
        if field == "price":
            result["price"] = _parse_price(raw_val)
        elif field == "volumes":
            result["volumes"] = _parse_volumes(raw_val)
        elif field in ("refillable", "is_new", "is_featured"):
            result[field] = _parse_bool(raw_val)
        else:
            result[field] = raw_val or None
    
    return result


def parse_csv_content(content: str) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse CSV string content into headers and rows."""
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    rows = list(reader)
    return headers, rows


def parse_xlsx_content(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse XLSX content bytes into headers and rows."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel support. pip install openpyxl")
    
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Empty Excel file")
    
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = []
    raw_data = []
    
    for row in rows_iter:
        if not raw_headers:
            raw_headers = [str(h) for h in row]
        else:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(raw_headers):
                    row_dict[raw_headers[i]] = str(val) if val is not None else ""
            raw_data.append(row_dict)
    
    wb.close()
    return raw_headers, raw_data


def process_import_rows(
    headers: List[str], rows: List[Dict[str, str]]
) -> Dict[str, Any]:
    """Process parsed rows into validated product data and errors."""
    imported = []
    errors = []
    
    for idx, row in enumerate(rows, start=2):  # row 1 = header
        product_data = _parse_row_to_product(row, headers)
        validation_errors = _validate_row(product_data)
        
        if validation_errors:
            for err in validation_errors:
                errors.append({
                    "row": idx,
                    "field": err["field"],
                    "message": err["message"],
                    "raw_data": {k: str(v) for k, v in row.items()},
                })
        else:
            imported.append(product_data)
    
    return {"imported": imported, "errors": errors}


class ProductImportService:
    """Service for importing products from CSV/Excel files."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def import_products(
        self, headers: List[str], rows: List[Dict[str, str]]
    ) -> Dict[str, Any]:
        """Parse, validate, and import products from file data."""
        result = process_import_rows(headers, rows)
        
        imported_count = 0
        for product_data in result["imported"]:
            try:
                from models.products import Products
                from services.products import ProductsService
                
                service = ProductsService(self.db)
                
                # Clean up None values and convert volumes string to list
                clean_data = {
                    k: v for k, v in product_data.items() if v is not None
                }
                
                # Convert volumes from string "0, 10, 20" to [0, 10, 20]
                if "volumes" in clean_data and isinstance(clean_data["volumes"], str):
                    try:
                        clean_data["volumes"] = [
                            int(x.strip())
                            for x in clean_data["volumes"].split(",")
                            if x.strip().isdigit()
                        ]
                    except (ValueError, AttributeError):
                        clean_data["volumes"] = []
                
                # Set defaults for boolean fields not present
                for bool_field in ["refillable", "is_new", "is_featured"]:
                    if bool_field not in clean_data:
                        clean_data[bool_field] = False
                
                # Check for duplicate name+brand
                existing = await service.get_by_field("name", clean_data.get("name", ""))
                # Skip duplicate check - just create
                
                await service.create(clean_data)
                imported_count += 1
            except Exception as e:
                result["errors"].append({
                    "row": "unknown",
                    "field": "all",
                    "message": str(e),
                    "raw_data": str(product_data),
                })
        
        return {
            "imported_count": imported_count,
            "errors": result["errors"],
            "total_processed": len(rows),
        }