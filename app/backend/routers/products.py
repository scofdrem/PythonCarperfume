import json
import logging
from typing import Any, List, Optional

from datetime import datetime, date

from dependencies.auth import get_admin_user
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from models.auth import User
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from core.database import get_db
from services.products import ProductsService
from services.product_import import (
    PRODUCT_FIELD_REGISTRY,
    parse_csv_content,
    parse_xlsx_content,
    process_import_rows,
    ProductImportService,
)

# Set up logging
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/entities/products", tags=["products"])


# ---------- Pydantic Schemas ----------
class ProductsData(BaseModel):
    """Entity data schema (for create/update)"""
    name: str
    brand: str
    price: float
    category: str = None
    volumes: str = None
    image: str = None
    description: str = None
    instagram_url: str = None
    refillable: bool = None
    is_new: bool = None
    is_featured: bool = None


class ProductsUpdateData(BaseModel):
    """Update entity data (partial updates allowed)"""
    name: Optional[str] = None
    brand: Optional[str] = None
    price: Optional[float] = None
    category: Optional[str] = None
    volumes: Optional[str] = None
    image: Optional[str] = None
    description: Optional[str] = None
    instagram_url: Optional[str] = None
    refillable: Optional[bool] = None
    is_new: Optional[bool] = None
    is_featured: Optional[bool] = None


class ProductsResponse(BaseModel):
    """Entity response schema"""
    id: int
    name: str
    brand: str
    price: Optional[float] = None
    category: Optional[str] = None
    volumes: Optional[str] = None
    image: Optional[str] = None
    description: Optional[str] = None
    instagram_url: Optional[str] = None
    refillable: Optional[bool] = None
    is_new: Optional[bool] = None
    is_featured: Optional[bool] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class ProductsListResponse(BaseModel):
    """List response schema"""
    items: List[ProductsResponse]
    total: int
    skip: int
    limit: int


class ProductsBatchCreateRequest(BaseModel):
    """Batch create request"""
    items: List[ProductsData]


class ProductsBatchUpdateItem(BaseModel):
    """Batch update item"""
    id: int
    updates: ProductsUpdateData


class ProductsBatchUpdateRequest(BaseModel):
    """Batch update request"""
    items: List[ProductsBatchUpdateItem]


class ProductsBatchDeleteRequest(BaseModel):
    """Batch delete request"""
    ids: List[int]


# ---------- Routes ----------
@router.get("", response_model=ProductsListResponse)
async def query_productss(
    query: str = Query(None, description="Query conditions (JSON string)"),
    sort: str = Query(None, description="Sort field (prefix with '-' for descending)"),
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(20, ge=1, le=2000, description="Max number of records to return"),
    fields: str = Query(None, description="Comma-separated list of fields to return"),
    db: AsyncSession = Depends(get_db),
):
    """Query productss with filtering, sorting, and pagination"""
    logger.debug(f"Querying productss: query={query}, sort={sort}, skip={skip}, limit={limit}, fields={fields}")
    
    service = ProductsService(db)
    try:
        # Parse query JSON if provided
        query_dict = None
        if query:
            try:
                query_dict = json.loads(query)
            except json.JSONDecodeError:
                raise HTTPException(status_code=400, detail="Invalid query JSON format")
        
        result = await service.get_list(
            skip=skip, 
            limit=limit,
            query_dict=query_dict,
            sort=sort,
        )
        logger.debug(f"Found {result['total']} productss")
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error querying productss: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.get("/all", response_model=ProductsListResponse)
async def query_productss_all(
    query: str = Query(None, description="Query conditions (JSON string)"),
    sort: str = Query(None, description="Sort field (prefix with '-' for descending)"),
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(20, ge=1, le=2000, description="Max number of records to return"),
    fields: str = Query(None, description="Comma-separated list of fields to return"),
    db: AsyncSession = Depends(get_db),
):
    # Query productss with filtering, sorting, and pagination without user limitation
    logger.debug(f"Querying productss: query={query}, sort={sort}, skip={skip}, limit={limit}, fields={fields}")

    service = ProductsService(db)
    try:
        # Parse query JSON if provided
        query_dict = None
        if query:
            try:
                query_dict = json.loads(query)
            except json.JSONDecodeError:
                raise HTTPException(status_code=400, detail="Invalid query JSON format")

        result = await service.get_list(
            skip=skip,
            limit=limit,
            query_dict=query_dict,
            sort=sort
        )
        logger.debug(f"Found {result['total']} productss")
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error querying productss: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.get("/{id}", response_model=ProductsResponse)
async def get_products(
    id: int,
    fields: str = Query(None, description="Comma-separated list of fields to return"),
    db: AsyncSession = Depends(get_db),
):
    """Get a single products by ID"""
    logger.debug(f"Fetching products with id: {id}, fields={fields}")
    
    service = ProductsService(db)
    try:
        result = await service.get_by_id(id)
        if not result:
            logger.warning(f"Products with id {id} not found")
            raise HTTPException(status_code=404, detail="Products not found")
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching products {id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.post("", response_model=ProductsResponse, status_code=201)
async def create_products(
    data: ProductsData,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    """Create a new products"""
    logger.debug(f"Creating new products with data: {data}")
    
    service = ProductsService(db)
    try:
        result = await service.create(data.model_dump())
        if not result:
            raise HTTPException(status_code=400, detail="Failed to create products")
        
        logger.info(f"Products created successfully with id: {result.id}")
        return result
    except ValueError as e:
        logger.error(f"Validation error creating products: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Error creating products: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.post("/batch", response_model=List[ProductsResponse], status_code=201)
async def create_productss_batch(
    request: ProductsBatchCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    """Create multiple productss in a single request"""
    logger.debug(f"Batch creating {len(request.items)} productss")
    
    service = ProductsService(db)
    results = []
    
    try:
        for item_data in request.items:
            result = await service.create(item_data.model_dump())
            if result:
                results.append(result)
        
        logger.info(f"Batch created {len(results)} productss successfully")
        return results
    except Exception as e:
        await db.rollback()
        logger.error(f"Error in batch create: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Batch create failed: {str(e)}")


@router.put("/batch", response_model=List[ProductsResponse])
async def update_productss_batch(
    request: ProductsBatchUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    """Update multiple productss in a single request"""
    logger.debug(f"Batch updating {len(request.items)} productss")
    
    service = ProductsService(db)
    results = []
    
    try:
        for item in request.items:
            # Only include non-None values for partial updates
            update_dict = {k: v for k, v in item.updates.model_dump().items() if v is not None}
            result = await service.update(item.id, update_dict)
            if result:
                results.append(result)
        
        logger.info(f"Batch updated {len(results)} productss successfully")
        return results
    except Exception as e:
        await db.rollback()
        logger.error(f"Error in batch update: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Batch update failed: {str(e)}")


@router.put("/{id}", response_model=ProductsResponse)
async def update_products(
    id: int,
    data: ProductsUpdateData,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    """Update an existing products"""
    logger.debug(f"Updating products {id} with data: {data}")

    service = ProductsService(db)
    try:
        # Only include non-None values for partial updates
        update_dict = {k: v for k, v in data.model_dump().items() if v is not None}
        result = await service.update(id, update_dict)
        if not result:
            logger.warning(f"Products with id {id} not found for update")
            raise HTTPException(status_code=404, detail="Products not found")
        
        logger.info(f"Products {id} updated successfully")
        return result
    except HTTPException:
        raise
    except ValueError as e:
        logger.error(f"Validation error updating products {id}: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Error updating products {id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.delete("/batch")
async def delete_productss_batch(
    request: ProductsBatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    """Delete multiple productss by their IDs"""
    logger.debug(f"Batch deleting {len(request.ids)} productss")
    
    service = ProductsService(db)
    deleted_count = 0
    
    try:
        for item_id in request.ids:
            success = await service.delete(item_id)
            if success:
                deleted_count += 1
        
        logger.info(f"Batch deleted {deleted_count} productss successfully")
        return {"message": f"Successfully deleted {deleted_count} productss", "deleted_count": deleted_count}
    except Exception as e:
        await db.rollback()
        logger.error(f"Error in batch delete: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Batch delete failed: {str(e)}")


@router.delete("/{id}")
async def delete_products(
    id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
):
    """Delete a single products by ID"""
    logger.debug(f"Deleting products with id: {id}")
    
    service = ProductsService(db)
    try:
        success = await service.delete(id)
        if not success:
            logger.warning(f"Products with id {id} not found for deletion")
            raise HTTPException(status_code=404, detail="Products not found")
        
        logger.info(f"Products {id} deleted successfully")
        return {"message": "Products deleted successfully", "id": id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting products {id}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")


@router.get("/template")
async def download_template() -> Any:
    """
    Download an Excel template for bulk product import.
    Dynamically generated from the Product model field registry.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Run: pip install openpyxl"
        )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Header row
    headers = list(PRODUCT_FIELD_REGISTRY.keys())
    for col_idx, field_name in enumerate(headers, start=1):
        meta = PRODUCT_FIELD_REGISTRY[field_name]
        cell = ws.cell(row=1, column=col_idx, value=field_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Highlight required fields with red
        if meta["required"]:
            cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    
    # Example row
    for col_idx, field_name in enumerate(headers, start=1):
        meta = PRODUCT_FIELD_REGISTRY[field_name]
        ws.cell(row=2, column=col_idx, value=meta["example"])
    
    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(headers[col_idx - 1]) + 4)
    
    # Add instructions
    ws.cell(row=4, column=1, value="Instructions:")
    ws.cell(row=4, column=1).font = Font(bold=True)
    ws.cell(row=5, column=1, value="  Red headers = required fields (name, brand, price)")
    ws.cell(row=6, column=1, value="  Volumes: comma-separated numbers, e.g. '0, 10, 20, 30'")
    ws.cell(row=7, column=1, value="  Booleans: TRUE/FALSE, Yes/No, 1/0")
    ws.cell(row=8, column=1, value="  Save as .xlsx or .csv and upload via Admin panel")
    
    # Write to buffer
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from starlette.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="product_import_template.xlsx"'}
    )


@router.post("/import")
async def import_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_user),
) -> Any:
    """
    Import products from an Excel (.xlsx, .xls) or CSV file.
    Returns count of imported products and any validation errors.
    """
    filename = (file.filename or "").lower()
    content = await file.read()
    
    try:
        if filename.endswith((".xlsx", ".xls")):
            headers, rows = parse_xlsx_content(content)
        elif filename.endswith(".csv"):
            headers, rows = parse_csv_content(content.decode("utf-8"))
        else:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file format. Use .xlsx, .xls, or .csv"
            )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    if not rows:
        raise HTTPException(status_code=400, detail="File contains no data rows")
    
    # Validate structure
    parsed_headers = [h.strip() for h in headers if h.strip()]
    if not parsed_headers:
        raise HTTPException(status_code=400, detail="No headers found in file")
    
    service = ProductImportService(db)
    result = await service.import_products(parsed_headers, rows)
    
    return {
        "imported_count": result["imported_count"],
        "total_processed": result["total_processed"],
        "errors": result["errors"],
        "success": len(result["errors"]) == 0,
    }
